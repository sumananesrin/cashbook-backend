from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db.models import Sum, Q, F, Value, CharField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from .models import Cashbook, Transaction, Member
from .serializers import CashbookSerializer, TransactionSerializer

import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from datetime import datetime

class ReportsViewSet(viewsets.GenericViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Get cashbooks user has access to"""
        user = self.request.user
        return Cashbook.objects.filter(
            Q(business__owner=user) | 
            Q(business__members__user=user)
        ).distinct()

    def list(self, request):
        """List all cashbooks with summary stats"""
        cashbooks = self.get_queryset()
        data = []
        
        for cashbook in cashbooks:
            # Calculate totals
            total_in = Transaction.objects.filter(cashbook=cashbook, type='IN').aggregate(Sum('amount'))['amount__sum'] or 0
            total_out = Transaction.objects.filter(cashbook=cashbook, type='OUT').aggregate(Sum('amount'))['amount__sum'] or 0
            net_balance = total_in - total_out
            
            last_txn = Transaction.objects.filter(cashbook=cashbook).order_by('-created_at').first()
            last_updated = last_txn.created_at if last_txn else cashbook.created_at

            data.append({
                'id': cashbook.id,
                'name': cashbook.name,
                'business_name': cashbook.business.name,
                'total_in': total_in,
                'total_out': total_out,
                'net_balance': net_balance,
                'last_updated': last_updated
            })
            
        return Response(data)

    def _get_report_data(self, cashbook):
        """Helper to get transactions with running balance"""
        transactions = Transaction.objects.filter(cashbook=cashbook).order_by('transaction_date', 'created_at')
        
        running_balance = 0
        txns_data = []
        
        total_in = 0
        total_out = 0

        for txn in transactions:
            if txn.type == 'IN':
                running_balance += txn.amount
                total_in += txn.amount
            else:
                running_balance -= txn.amount
                total_out += txn.amount
            
            txns_data.append({
                'date': txn.transaction_date,
                'time': txn.transaction_time.strftime('%H:%M') if txn.transaction_time else '',
                'type': txn.type,
                'party': txn.party.name if txn.party else '-',
                'category': txn.category.name if txn.category else '-',
                'payment_mode': txn.payment_mode.name if txn.payment_mode else '-',
                'remark': txn.remark,
                'amount_in': txn.amount if txn.type == 'IN' else 0,
                'amount_out': txn.amount if txn.type == 'OUT' else 0,
                'running_balance': running_balance
            })
            
        return {
            'cashbook_name': cashbook.name,
            'transactions': txns_data,
            'total_in': total_in,
            'total_out': total_out,
            'net_balance': total_in - total_out
        }

    def retrieve(self, request, pk=None):
        """Get detailed report for a specific cashbook"""
        cashbook = self.get_object() # Standard get_object checks permission if using standard mixins, but custom get_queryset handles filtering
        if not cashbook:
             return Response({'error': 'Cashbook not found'}, status=404)

        report_data = self._get_report_data(cashbook)
        # Reverse transactions for display (newest first) but keep calculation order correct
        report_data['transactions'].reverse()
        
        return Response(report_data)

    @action(detail=True, methods=['get'], url_path='export_excel')
    def export_excel(self, request, pk=None):
        cashbook = self.get_object()
        data = self._get_report_data(cashbook)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{cashbook.name} Report"
        
        # Headers
        headers = ['Date', 'Type', 'Party', 'Category', 'Payment Mode', 'Remark', 'Cash In', 'Cash Out', 'Balance']
        ws.append(headers)
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Data
        for txn in data['transactions']:
            ws.append([
                txn['date'],
                txn['type'],
                txn['party'],
                txn['category'],
                txn['payment_mode'],
                txn['remark'],
                txn['amount_in'] if txn['amount_in'] > 0 else '',
                txn['amount_out'] if txn['amount_out'] > 0 else '',
                txn['running_balance']
            ])
            
        # Summary at bottom
        last_row = len(data['transactions']) + 3
        ws.cell(row=last_row, column=6, value="TOTALS").font = openpyxl.styles.Font(bold=True)
        ws.cell(row=last_row, column=7, value=data['total_in']).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=last_row, column=8, value=data['total_out']).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=last_row, column=9, value=data['net_balance']).font = openpyxl.styles.Font(bold=True)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{cashbook.name}_report.xlsx"'
        
        wb.save(response)
        return response

    @action(detail=True, methods=['get'], url_path='export_pdf')
    def export_pdf(self, request, pk=None):
        cashbook = self.get_object()
        data = self._get_report_data(cashbook)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        elements.append(Paragraph(f"Report: {cashbook.name}", styles['Title']))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Summary
        summary_data = [
            ['Total In', 'Total Out', 'Net Balance'],
            [f"{data['total_in']:.2f}", f"{data['total_out']:.2f}", f"{data['net_balance']:.2f}"]
        ]
        t_summary = Table(summary_data, colWidths=[100, 100, 100])
        t_summary.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t_summary)
        elements.append(Spacer(1, 20))
        
        # Transactions Table
        table_data = [['Date', 'Type', 'Party', 'Category', 'Mode', 'Remark', 'In', 'Out', 'Balance']]
        
        for txn in data['transactions']:
            table_data.append([
                str(txn['date']),
                txn['type'],
                txn['party'],
                txn['category'],
                txn['payment_mode'],
                txn['remark'][:20], # Truncate remark for PDF
                f"{txn['amount_in']:.2f}" if txn['amount_in'] > 0 else '',
                f"{txn['amount_out']:.2f}" if txn['amount_out'] > 0 else '',
                f"{txn['running_balance']:.2f}"
            ])
            
        t = Table(table_data, colWidths=[60, 40, 80, 80, 60, 100, 60, 60, 70])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{cashbook.name}_report.pdf"'
        return response
